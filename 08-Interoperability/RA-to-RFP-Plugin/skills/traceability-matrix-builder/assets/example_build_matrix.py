from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLUE="1F4E79"; HDR="2E75B6"; LB="D6E4F0"; GREY="F2F2F2"; GREEN="E2EFDA"; AMBER="FFF2CC"
thin=Side(style="thin",color="BFBFBF")
bd=Border(left=thin,right=thin,top=thin,bottom=thin)
hfont=Font(name="Arial",bold=True,color="FFFFFF",size=10)
tfont=Font(name="Arial",size=10)
tbold=Font(name="Arial",size=10,bold=True)
wrap=Alignment(wrap_text=True,vertical="top")
ctr=Alignment(horizontal="center",vertical="center",wrap_text=True)

wb=Workbook()

def style_header(ws,ncol,row=1,fill=HDR):
    for c in range(1,ncol+1):
        cell=ws.cell(row=row,column=c)
        cell.font=hfont; cell.fill=PatternFill("solid",fgColor=fill)
        cell.alignment=ctr; cell.border=bd

def write(ws,rows,widths,freeze="A2",header_fill=HDR,prio_col=None):
    for r,row in enumerate(rows,1):
        for c,val in enumerate(row,1):
            cell=ws.cell(row=r,column=c,value=val)
            cell.border=bd
            if r==1:
                pass
            else:
                cell.font=tfont; cell.alignment=wrap
                if c==1: cell.font=tbold
    style_header(ws,len(widths),1,header_fill)
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes=freeze
    # priority colour
    if prio_col:
        for r in range(2,len(rows)+1):
            v=str(ws.cell(row=r,column=prio_col).value or "")
            if v.startswith("P1"): ws.cell(row=r,column=prio_col).fill=PatternFill("solid",fgColor="FCE4D6")
            elif v.startswith("P2"): ws.cell(row=r,column=prio_col).fill=PatternFill("solid",fgColor="FFF2CC")
            elif v.startswith("P3"): ws.cell(row=r,column=prio_col).fill=PatternFill("solid",fgColor=GREEN)
    for r in range(1,len(rows)+1):
        ws.row_dimensions[r].height=None

# ---------------- Sheet 1: Cover / README ----------------
ws0=wb.active; ws0.title="README"
readme=[
 ["Requirements Traceability Matrix — Government Interoperability Platform (GIP)"],
 ["GICTA — WARDIP (World Bank) — companion to the RFP v1.0 (DRAFT, June 2026)"],
 [""],
 ["Purpose"],
 ["Maps every mandatory requirement in the RFP (Annex A - Consolidated Interoperability Requirements) and the"],
 ["National Enterprise Architecture (NEA) to the RFP clause that procures it and to an acceptance test the"],
 ["independent IV&V/QA agent uses to verify delivery before GICTA accepts and pays. Doubles as the IV&V test basis."],
 [""],
 ["Sheets"],
 ["1. README — this sheet"],
 ["2. Requirements Traceability — RA + NEA mandatory requirements -> RFP clause -> acceptance test"],
 ["3. National Integration Map — the 26 NEA integration points -> RFP mapping -> acceptance test"],
 ["4. Verification Log — blank columns for IV&V status, evidence and date (to be filled during delivery)"],
 [""],
 ["How to use"],
 ["- 'RFP clause' points to the section of the RFP that imposes the requirement on the Supplier."],
 ["- 'Acceptance test / IV&V check' is the objective test that must pass for the deliverable to be accepted."],
 ["- 'Verification status' is left blank; the IV&V agent records Pass/Fail, evidence and date during delivery."],
 ["- Granularity: the 'Annex A REQ Trace' sheet lists every REQ-ID with its acceptance basis and phase. Bespoke per-REQ-ID test"],
 ["  specs are baselined by the Supplier with the IV&V agent at the Inception gate (an inception deliverable), then recorded here."],
 [""],
 ["Sources: RFP Annex A (consolidated interoperability requirements); GovStack specifications; National Enterprise Architecture."],
]
for r,row in enumerate(readme,1):
    cell=ws0.cell(row=r,column=1,value=row[0])
    if r==1: cell.font=Font(name="Arial",bold=True,size=14,color=BLUE)
    elif r==2: cell.font=Font(name="Arial",italic=True,size=10,color="666666")
    elif row[0] in ("Purpose","Sheets","How to use"): cell.font=Font(name="Arial",bold=True,size=11,color=HDR)
    else: cell.font=Font(name="Arial",size=10)
ws0.column_dimensions["A"].width=110

# ---------------- Sheet 2: Requirements Traceability ----------------
ws=wb.create_sheet("Requirements Traceability")
hdr=["ID","Source","Reference","Mandatory requirement","RFP clause","Stream","Acceptance test / IV&V check","Phase","Priority"]
R=[hdr]
def row(i,src,ref,req,tor,stream,test,phase,pri): R.append([i,src,ref,req,tor,stream,test,phase,pri])

# RA-derived
row("R-01","RA","§4.4","Regulator/operator split enforced; GICTA = regulator/owner & eventual operator; Supplier builds & operates then transfers","§4","Gov","Governance charter & RACI delivered separating rule-making, operation and compliance; GICTA sign-off","Foundation","P1")
row("R-02","RA","§4.7, §13","Compliance verified before procurement and before go-live (no production without compliance)","§4, §8","Gov","Onboarding/compliance checkpoints documented; no pilot service goes live without a passed compliance gate","Pilot","P1")
row("R-03","RA","§3.1–3.5","Interoperability legal framework & data-protection instruments drafted (two-agreement model)","§3.2","Legal","Legal instruments + DSA framework delivered and GICTA-approved; cover lawful basis & data protection","Foundation","P1")
row("R-04","RA","§5.1","Service ownership model established; every shared service has a named owner","§3.2","Gov","Service-ownership model published; each pilot service has a documented owner","Foundation","P2")
row("R-05","RA","§6.2","Base registries with authoritative-source resolution","§3.2","Sem","Authoritative-source resolution documented for pilot data; consumers read from the authoritative source","Foundation","P1")
row("R-06","RA","§6.3–6.4","Metadata management & semantic standards applied","§3.2","Sem","Metadata catalogue populated for pilot services; semantic standards applied & verified","Foundation","P2")
row("R-07","RA","§7.1","GIP built on GovStack Information Mediator BB (X-Road): Security Server, Service Access, Pub/Sub, API gateway","§3.1","Tech","GIP components deployed; federated/distributed model demonstrated on the pilot","Foundation","P1")
row("R-08","RA","§7.2; NEA §4.2.2","Technical standards implemented (OpenAPI 3.1, REST/JSON; domain extensions only where specified)","§7","Tech","Every pilot API registered with a valid OpenAPI 3.1 spec; no new SOAP/XML integrations","Pilot","P1")
row("R-09","RA","§7.3","Openness; GICTA owns/controls source code; adapter pattern mandated (no vendor lock-in)","§7","Tech","Source code & build artefacts handed to GICTA under licence; connectors isolated via adapters; portability shown","Transfer","P1")
row("R-10","RA","§7.5","National CFR baseline mandatory (GovStack Cross-Functional Specifications) in design, delivery, test, prod approval","§7","Tech","Each pilot service passes the CFR checklist; production approval gated on CFR conformance","Pilot","P1")
row("R-11","RA","§7.7","Trust services & PKI (OAuth 2.0 / OIDC / mutual TLS)","§3.1, §7","Tech","PKI operational; mutual-TLS + OIDC enforced on pilot exchanges; certificate lifecycle documented","Foundation","P1")
row("R-12","RA","§7.8","Service registry & API catalogue","§3.1","Tech","Registry/catalogue live; all pilot services discoverable with current specs","Pilot","P1")
row("R-13","RA","§7.9","Event management architecture (async pub/sub)","§3.1","Tech","Pub/Sub layer demonstrated by the event-driven pilot flow (INT-16)","Pilot","P1")
row("R-14","RA","§7.10","Security architecture & minimum security baseline; SOC monitoring of platform traffic","§7","Tech","Security baseline evidenced (pen-test/scan); GIP traffic visible to the national SOC","Pilot","P1")
row("R-15","RA","§7.11","Exchange-layer access-logging & consent built into the GIP; separate national systems (GovPay, IAM, portal, notification) integrated via standard interfaces (reuse-first), not rebuilt","§3.1","Tech","Access-logging operational before personal data flows; no parallel payment/identity/portal/notification built; integrations use standard interfaces","Pilot","P1")
row("R-16","RA","§7.12","AI-readiness / AI-operable services","§7","Tech","Pilot services expose machine-actionable specs consistent with RA §7.12","Pilot","P3")
row("R-17","RA","§7.13","Five-stage onboarding lifecycle; gates not bypassed; test/prod separated","§6, §8","Tech","Each pilot agency onboarded through the five gates; evidence of completed test exchange before go-live","Pilot","P1")
row("R-18","RA","§7.14","Legacy transition & exception handling; SOAP-to-REST adapters at boundaries; exceptions time-bound & registered","§3.1","Tech","Any exception recorded in the §4.7 register with a regularisation deadline; adapters at legacy boundaries","Pilot","P2")
row("R-19","RA","§7.16, §8.3","Deployment architecture; environment separation; resilience & BC/DR","§3.1","Tech","Separate prod/non-prod; BC/DR plan tested; recovery objectives met in a failover test","Foundation","P1")
row("R-20","RA","§8.1","Hosting on GICTA-provided NDC / sovereign G-Cloud","§13","Infra","GIP deployed on GICTA-provided hosting; sovereignty/data-residency confirmed","Foundation","P1")
row("R-21","RA","§10.1–10.2","Capability-building (Academy) & priority-skills tracks taught against real services","§9","Cap","Academy tracks delivered against pilot services; attendance & competency records","Operate","P1")
row("R-22","RA","§12","Sustainability/funding model for post-transfer operation","§3.3, §8","Cap","Sustainability model delivered; feeds the WARDIP->MTEF transition (2029 budget line identified)","Transfer","P1")
# NEA-derived
row("R-23","NEA","§4.1","GIP-centred topology; no point-to-point cross-ministry integration in target state","§7","Tech","Pilot flows route through the GIP; no new point-to-point links introduced","Pilot","P1")
row("R-24","NEA","§4.2","Two-tier mediation standard (national GIP = Tier 1; sector mediators = Tier 2)","§7","Tech","Pilot conforms to two-tier pattern; GIP handles routing/security, not business logic","Pilot","P1")
row("R-25","NEA","§4.2.2","OpenAPI 3.1/REST-JSON baseline + OAuth2/OIDC/mTLS; SOAP only via adapters","§7","Tech","Conformance test of pilot APIs to the standard; auth model verified","Pilot","P1")
row("R-26","NEA","§5.3","Four identity master keys (NIN, TIN, MPI, MDA codes) usable for cross-ministry matching","§3.2","Sem","TIN verification service exposed via GIP for the revenue pilot; master-key approach documented","Foundation","P1")
row("R-27","NEA","§4.3.4","INT-07 Chart-of-Accounts mapping ('Integration Point Zero') version-controlled & published","§5","Sem","Machine-readable CoA mapping (semantic versioning) published to the registry; reconciles on real rows","Foundation","P1")
row("R-28","NEA","§5.4","Five-tier national data-classification framework; GIP mediation rules per tier; DSA-gated","§7","Sem","Classification applied to pilot data; tier-appropriate handling & DSA enforced; crosswalk ready for NDGB","Foundation","P1")
row("R-29","NEA","§5.5","Data Sharing Agreement framework (two-agreement model)","§3.2","Legal","DSA template delivered; signed DSAs in place for each pilot flow before go-live","Pilot","P1")
row("R-30","NEA","§4.4.3","'Design for federation' — sector mediators reserve GIP endpoints from inception","§6","Tech","Pilot designs include reserved GIP federation endpoints; no isolated-silo design","Foundation","P2")
row("R-31","NEA","§4.1, §4.3.2","Phase-1 Integration Paradox: interim bilateral connections must be GIP-compatible & migratable","§3.1, §17","Tech","Any interim connection uses GIP-compatible specs (OpenAPI 3.1, mTLS); migration is a deployment change","Pilot","P1")
# Re-label RA source as ToR Annex A (RA is not cited as an authoritative source)
import re as _re
def _ra_layer(ref):
    m=_re.search(r'§\s*(\d+)',ref); ch=m.group(1) if m else '7'
    L={'1':'A','2':'A','3':'B','4':'C','5':'D','6':'E','7':'F','8':'G'}.get(ch,'H')
    return _re.sub(r'§\s*\d+(\.\d+)*(\s*[–-]\s*\d+(\.\d+)*)?','Annex A §'+L,ref,count=1)
for rr in R[1:]:
    if rr[1]=='RA':
        rr[1]='Annex A'; rr[2]=_ra_layer(rr[2])
# write
ws.delete_rows(1, ws.max_row)
for rr in R:
    ws.append(rr)
write(ws,R,[7,8,12,42,10,7,46,11,9],freeze="A2",prio_col=9)

# ---------------- Sheet 3: National Integration Map ----------------
ws2=wb.create_sheet("National Integration Map")
ih=["Ref","Integration point","Priority","Target mechanism","In pilot?","RFP mapping","Acceptance test / IV&V check"]
IM=[ih]
def im(ref,pt,pri,mech,pilot,tor,test): IM.append([ref,pt,pri,mech,pilot,tor,test])
im("INT-01","IFMIS <- GRA (Revenue Data)","P1-Critical","Batch->API(GIP)->Event","Follow-on","Annex B","—")
im("INT-02","GovPay <-> IFMIS","P1-Enhance","Direct API","Follow-on","Annex B","—")
im("INT-03","GovPay <-> GRA GamTax Net","P1-Critical","Direct API","Follow-on","Annex B","—")
im("INT-04","GovPay <-> GRA ASYCUDA","P1-Critical","Direct API","Follow-on","Annex B","—")
im("INT-05","GRA -> MoFEA (Tax Arrears)","P2","GIP","Follow-on","Annex B","—")
im("INT-06","GRA -> MoFEA (Exemption Register)","P2","GIP","Follow-on","Annex B","—")
im("INT-07","CoA Revenue Mapping (Integration Point Zero)","P1-Critical","Batch / registry","YES","§5; R-27","Versioned CoA mapping published to registry; reconciles on real GRA/IFMIS rows")
im("INT-08","IFMIS -> GIP (Financial APIs)","P1","GIP","YES","§5","IFMIS financial query served via GIP with valid OpenAPI 3.1 spec & mTLS")
im("INT-09","GovPay <-> MyGov","P2","GIP","Follow-on","Annex B","—")
im("INT-10","MoFEA -> Open Data","P3","GIP","Follow-on","Annex B","—")
im("INT-11","G-Cloud <- IFMIS (hosting)","P2-3","Infra migration","Follow-on","Annex B","—")
im("INT-12","IFMIS <- MoH (Health Budget)","P2","GIP","Follow-on","Annex B","—")
im("INT-13","GovPay <- MoH (Health Fees)","P2","GIP","Follow-on","Annex B","—")
im("INT-14","GRA <-> GIP (Tax/Customs APIs)","P1","GIP","YES","§5","GRA tax/customs query served via GIP; sync response verified end-to-end")
im("INT-15","GRA <-> National IAM (TIN<->NIN)","P2","GIP","Follow-on","Annex B","—")
im("INT-16","GRA <- GIEPA (Business Registration)","P2","GIP (event-driven)","YES","§5; R-13","Business-registration event triggers automatic TIN assignment via Pub/Sub")
im("INT-17","GRA <- NDAS (Address Data)","P2","GIP","Follow-on","Annex B","—")
im("INT-18","eCRVS <-> MPI (via HIE)","P1-Critical","Direct API (OpenHIM)","Follow-on","Annex B","Deferred: Tier-2 OpenHIM/FHIR")
im("INT-19","eCRVS <-> National Identity Broker","P2","GIP","Follow-on","Annex B","—")
im("INT-20","DHIS2 -> G-Cloud","P2","Infra migration","Follow-on","Annex B","—")
im("INT-21","MoH Systems <-> GIP","P2","GIP","Follow-on","Annex B","—")
im("INT-22","GRA <- eCRVS (Death Registration)","P3","GIP","Follow-on","Annex B","—")
im("INT-23","GIP <-> All Tier-1 Systems","P1-Foundational","X-Road GIP backbone","YES (platform)","§3.1; R-07","GIP operational as the backbone; all pilot flows transit it")
im("INT-24","National IAM <-> All Systems","P2","GIP","Follow-on","Annex B","—")
im("INT-25","NCSC/SOC <-> All DPI Platforms","P1","GIP monitoring","YES (security)","§7; R-14","GIP traffic monitored by the national SOC; alerts demonstrated")
im("INT-26","HRMIS <-> GIP","P2","GIP","Follow-on","Annex B","—")
for rr in IM: ws2.append(rr)
write(ws2,IM,[8,42,13,20,13,12,52],freeze="A2",prio_col=3)

# ---------------- Sheet 4: Verification Log ----------------
ws3=wb.create_sheet("Verification Log")
vh=["Requirement ID","RFP clause","Verification status (Pass/Fail/Partial)","Evidence reference","Verified by (IV&V)","Date","Notes"]
V=[vh]
for rr in R[1:]:
    V.append([rr[0],rr[4],"","","","",""])
for rr in V: ws3.append(rr)
write(ws3,V,[14,10,28,22,18,12,40],freeze="A2")
for r in range(2,len(V)+1):
    ws3.cell(row=r,column=3).fill=PatternFill("solid",fgColor=AMBER)

# ---------------- Sheet 5: Annex A REQ-ID line-by-line traceability ----------------
import json
areas=json.load(open("authored_reqs.json"))
def phase_of(acc):
    a=acc.lower()
    if "foundation" in a: return "Foundation"
    if "transfer" in a: return "Transfer"
    if "pilot" in a or "go-live" in a: return "Pilot"
    return "All phases"
ws4=wb.create_sheet("Annex A REQ Trace")
qh=["REQ ID","Layer","Requirement area","Requirement (Supplier obligation)","Applicable standard","Acceptance basis","Phase","Verification status"]
Q=[qh]
for L in ["A","B","C","D","E","F","G","H"]:
    n=0
    for a in [x for x in areas if x["L"]==L]:
        for req in a["reqs"]:
            n+=1
            rid="REQ-%s-%02d"%(L,n)
            Q.append([rid,L,a["title"],req,a.get("std","") or "—",a["accept"],phase_of(a["accept"]),""])
for rr in Q: ws4.append(rr)
write(ws4,Q,[10,6,24,52,22,30,11,16],freeze="A2")
# colour status column
for r in range(2,len(Q)+1):
    ws4.cell(row=r,column=8).fill=PatternFill("solid",fgColor=AMBER)

# update README sheet list to mention sheet 5
ws0.cell(row=61,column=1,value="4. National Integration Map / Verification Log / Annex A REQ Trace — see tabs")

wb.save("Requirements_Traceability_Matrix_GIP.xlsx")
print("saved", ws.max_row-1, "theme reqs;", ws2.max_row-1, "integration points;", len(Q)-1, "REQ-ID rows")
